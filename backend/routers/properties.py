"""
KWAC OS -- Properties Router
"""
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from database import get_db
from auth import require_role

router = APIRouter()


def _prop_row(row) -> dict:
    return {
        "id": str(row["id"]),
        "ilist_code": row["ilist_code"],
        "status": row["status"],
        "transaction_type": row["transaction_type"],
        "address": row["address"],
        "area": row["area"],
        "municipality": row["municipality"],
        "property_type": row["property_type"],
        "sqm": float(row["sqm"]) if row["sqm"] else None,
        "floor": row["floor"],
        "bedrooms": row["bedrooms"],
        "price_asking": float(row["price_asking"]) if row["price_asking"] else None,
        "lat": float(row["lat"]) if row["lat"] else None,
        "lng": float(row["lng"]) if row["lng"] else None,
        "listing_type": row["listing_type"],
        "description": row["description"],
        "internal_notes": row.get("internal_notes"),
        "agent_id": str(row["agent_id"]) if row["agent_id"] else None,
        "agent_name": row["agent_name"],
        "agent_phone": row["agent_phone"],
        "agent_email": row["agent_email"],
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
    }

_SEL = (
    "SELECT p.id, p.ilist_code, p.status, p.transaction_type, "
    "p.address, p.area, p.municipality, p.property_type, "
    "p.sqm, p.floor, p.bedrooms, p.price_asking, "
    "p.lat, p.lng, p.listing_type, p.description, p.internal_notes, p.created_at, "
    "u.id AS agent_id, u.full_name AS agent_name, "
    "a.phone AS agent_phone, u.email AS agent_email "
    "FROM properties p "
    "LEFT JOIN users u ON u.id = p.agent_id "
    "LEFT JOIN agents a ON a.id = p.agent_id"
)


@router.get("/")
async def list_properties(
    status: Optional[str] = "active",
    area: Optional[str] = None,
    transaction_type: Optional[str] = None,
    q: Optional[str] = None,
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    conds, params = [], {}
    if status: conds.append("p.status = :status"); params["status"] = status
    if area: conds.append("p.area ILIKE :area"); params["area"] = "%" + area + "%"
    if transaction_type: conds.append("p.transaction_type = :tt"); params["tt"] = transaction_type
    if q:
        conds.append("(p.ilist_code ILIKE :q OR p.address ILIKE :q OR p.area ILIKE :q)")
        params["q"] = "%" + q + "%"
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    r = await db.execute(text(_SEL + " " + where + " ORDER BY p.created_at DESC LIMIT 300"), params)
    return [_prop_row(row) for row in r.mappings()]


@router.get("/by-code/{code}")
async def get_by_code(
    code: str,
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(text(_SEL + " WHERE p.ilist_code = :code LIMIT 1"), {"code": code})
    row = r.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Akinito den vrethike")
    return _prop_row(row)


class PropertyCreate(BaseModel):
    address: str
    area: Optional[str] = None
    municipality: Optional[str] = None
    transaction_type: str = "sale"
    property_type: Optional[str] = None
    sqm: Optional[float] = None
    floor: Optional[int] = None
    bedrooms: Optional[int] = None
    price_asking: Optional[float] = None
    status: str = "meeting"
    internal_notes: Optional[str] = None


@router.post("/")
async def create_property(
    body: PropertyCreate,
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    if body.status not in ("active","meeting","sold","rented","withdrawn"):
        raise HTTPException(status_code=400, detail="Invalid status")
    r = await db.execute(text(
        "INSERT INTO properties (agent_id, address, area, municipality, transaction_type, "
        "property_type, sqm, floor, bedrooms, price_asking, status, internal_notes) "
        "VALUES (:uid,:addr,:area,:muni,:tt,:ptype,:sqm,:floor,:beds,:price,:status,:notes) "
        "RETURNING id"
    ), {"uid": user["id"], "addr": body.address, "area": body.area, "muni": body.municipality,
        "tt": body.transaction_type, "ptype": body.property_type, "sqm": body.sqm,
        "floor": body.floor, "beds": body.bedrooms, "price": body.price_asking,
        "status": body.status, "notes": body.internal_notes})
    await db.commit()
    return {"ok": True, "id": str(r.scalar())}


class PropertyUpdate(BaseModel):
    status: Optional[str] = None
    price_asking: Optional[float] = None
    internal_notes: Optional[str] = None


@router.put("/{property_id}")
async def update_property(
    property_id: str, body: PropertyUpdate,
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    upd, params = [], {"id": property_id}
    if body.status is not None: upd.append("status=:status"); params["status"] = body.status
    if body.price_asking is not None: upd.append("price_asking=:price"); params["price"] = body.price_asking
    if body.internal_notes is not None: upd.append("internal_notes=:notes"); params["notes"] = body.internal_notes
    if upd:
        await db.execute(text("UPDATE properties SET " + ", ".join(upd) + ", updated_at=NOW() WHERE id=:id"), params)
        await db.commit()
    return {"ok": True}


# ── iList Excel Import ────────────────────────────────────────
@router.post("/import-ilist")
async def import_ilist(
    file: UploadFile = File(...),
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Apodekta mono .xlsx / .xls arxeia")
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=400, detail="Sfalma: " + str(e))

    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

    def find_col(*candidates):
        for c in candidates:
            for i, h in enumerate(headers):
                if c in h:
                    return i
        return None

    col_code  = find_col("kod", "code", "ar.")
    col_addr  = find_col("dieuth", "address", "odos")
    col_area  = find_col("periox", "area", "synoik", "topoth")
    col_muni  = find_col("dim", "munic", "pol")
    col_ptype = find_col("typos", "type", "eidos")
    col_tt    = find_col("pol", "enoik", "transaction", "kateg")
    col_sqm   = find_col("t.m", "sqm", "embad", "tm ")
    col_floor = find_col("orof", "floor")
    col_beds  = find_col("ypnod", "bed", "domat")
    col_price = find_col("tim", "price", "axia")
    col_stat  = find_col("katast", "status", "diathesim")
    col_agent = find_col("mesit", "agent", "ypeuth", "prakt", "symb")
    col_lt    = find_col("apokleis", "listing", "anath")

    ar = await db.execute(text("SELECT id, full_name FROM users WHERE is_active=TRUE"))
    agent_map = {row["full_name"].lower(): str(row["id"]) for row in ar.mappings()}

    def cell_val(row_vals, i):
        if i is None or i >= len(row_vals):
            return None
        v = row_vals[i]
        return str(v).strip() if v is not None else None

    def to_float(s):
        if not s: return None
        try: return float(s.replace(",", "").replace(".", "").replace(" ", "").replace("\u20ac", "")) or None
        except: return None

    def to_float_sqm(s):
        if not s: return None
        try: return float(s.replace(",", ".").replace(" ", "")) or None
        except: return None

    imported = updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        cv = lambda i: cell_val(row, i)

        code    = cv(col_code)
        address = cv(col_addr) or "Agnosi dieuthynsi"
        area    = cv(col_area)
        muni    = cv(col_muni)
        ptype   = cv(col_ptype)
        sqm     = to_float_sqm(cv(col_sqm))
        price   = to_float(cv(col_price))
        try: floor = int(str(cv(col_floor) or "0").split(".")[0])
        except: floor = None
        try: beds = int(str(cv(col_beds) or "0").split(".")[0])
        except: beds = None

        tt_raw = (cv(col_tt) or "").lower()
        tt = "rental" if any(x in tt_raw for x in ["enoik", "rent"]) else "sale"

        st_raw = (cv(col_stat) or "").lower()
        if any(x in st_raw for x in ["polithike", "sold"]): status = "sold"
        elif any(x in st_raw for x in ["enikiaste", "rented"]): status = "rented"
        elif any(x in st_raw for x in ["aposyrthe", "withdr"]): status = "withdrawn"
        else: status = "active"

        lt_raw = (cv(col_lt) or "").lower()
        listing_type = "exclusive" if "apokleis" in lt_raw else ("simple" if "apl" in lt_raw else None)

        agent_name = (cv(col_agent) or "").lower()
        agent_id = user["id"]
        if agent_name:
            for aname, aid in agent_map.items():
                if agent_name in aname or aname in agent_name:
                    agent_id = aid; break

        params = {"code": code, "addr": address, "area": area, "muni": muni, "tt": tt,
                  "ptype": ptype, "sqm": sqm, "floor": floor, "beds": beds, "price": price,
                  "status": status, "agent": agent_id, "lt": listing_type}

        if code:
            ex = await db.execute(text("SELECT id FROM properties WHERE ilist_code=:code"), {"code": code})
            if ex.first():
                await db.execute(text("""
                    UPDATE properties SET
                        address=:addr, area=:area, municipality=:muni, transaction_type=:tt,
                        property_type=:ptype, sqm=:sqm, floor=:floor, bedrooms=:beds,
                        price_asking=:price, status=:status, agent_id=:agent,
                        listing_type=:lt, updated_at=NOW()
                    WHERE ilist_code=:code
                """), params)
                updated += 1
                continue

        await db.execute(text("""
            INSERT INTO properties
                (ilist_code, address, area, municipality, transaction_type, property_type,
                 sqm, floor, bedrooms, price_asking, status, agent_id, listing_type)
            VALUES
                (:code, :addr, :area, :muni, :tt, :ptype,
                 :sqm, :floor, :beds, :price, :status, :agent, :lt)
        """), params)
        imported += 1

    await db.commit()
    return {"ok": True, "imported": imported, "updated": updated}


# ── Meeting Ακινήτων — Valuations ─────────────────────────────
class ValuationIn(BaseModel):
    estimated_price: Optional[float] = None
    comment: Optional[str] = None


@router.get("/{property_id}/valuations")
async def get_valuations(
    property_id: str,
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(text("""
        SELECT pv.id, pv.agent_id, u.full_name AS agent_name,
               pv.estimated_price, pv.comment, pv.created_at
        FROM property_valuations pv
        LEFT JOIN users u ON u.id = pv.agent_id
        WHERE pv.property_id = :pid
        ORDER BY pv.created_at
    """), {"pid": property_id})
    rows = r.mappings().all()
    return [
        {
            "id": str(row["id"]),
            "agent_id": str(row["agent_id"]),
            "agent_name": row["agent_name"],
            "estimated_price": float(row["estimated_price"]) if row["estimated_price"] else None,
            "comment": row["comment"],
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
        }
        for row in rows
    ]


@router.post("/{property_id}/valuations")
async def submit_valuation(
    property_id: str,
    body: ValuationIn,
    user=Depends(require_role("agent", "ceo", "admin")),
    db: AsyncSession = Depends(get_db),
):
    await db.execute(text("""
        INSERT INTO property_valuations (property_id, agent_id, estimated_price, comment)
        VALUES (:pid, :uid, :price, :comment)
        ON CONFLICT (property_id, agent_id) DO UPDATE SET
            estimated_price=:price, comment=:comment
    """), {"pid": property_id, "uid": user["id"],
          "price": body.estimated_price, "comment": body.comment})
    await db.commit()
    return {"ok": True}
